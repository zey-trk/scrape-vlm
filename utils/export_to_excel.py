"""
Excel Exporter - Exports RAG chunks to Excel with comprehensive formatting
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict


class ExcelExporter:
    """Exports RAG chunks to Excel with multiple sheets and formatting"""
    
    def __init__(self, chunks: List[Dict], stats: Dict, output_file: str = "rag_chunks_analysis.xlsx"):
        """
        Initialize the Excel exporter
        
        Args:
            chunks: List of all chunks from all strategies
            stats: Statistics dictionary from RAG chunker
            output_file: Output Excel file path
        """
        self.chunks = chunks
        self.stats = stats
        self.output_file = output_file
        
    def create_summary_sheet(self) -> pd.DataFrame:
        """Create summary statistics sheet"""
        summary_data = []
        
        for strategy, stat in self.stats.items():
            summary_data.append({
                'Strategy': strategy,
                'Total Chunks': stat['total_chunks'],
                'Avg Characters': round(stat['avg_chars'], 1),
                'Avg Words': round(stat['avg_words'], 1),
                'Total Characters': stat['total_chars'],
                'Total Words': stat['total_words'],
            })
        
        return pd.DataFrame(summary_data)
    
    def create_all_chunks_sheet(self) -> pd.DataFrame:
        """Create sheet with all chunks"""
        df_data = []
        
        for chunk in self.chunks:
            df_data.append({
                'Chunk ID': chunk['chunk_id'],
                'Strategy': chunk['strategy'],
                'Chunk Size Param': chunk['chunk_size_param'],
                'Overlap Param': chunk['overlap_param'],
                'Source URL': chunk['source_url'],
                'Page Title': chunk['page_title'],
                'Position in Doc': chunk['position_in_doc'],
                'Character Count': chunk['char_count'],
                'Word Count': chunk['word_count'],
                'Chunk Text': chunk['chunk_text'],
            })
        
        return pd.DataFrame(df_data)
    
    def create_strategy_sheet(self, strategy_name: str, chunk_size: int, overlap: int) -> pd.DataFrame:
        """Create sheet for specific strategy"""
        filtered_chunks = [
            c for c in self.chunks 
            if c['strategy'] == strategy_name 
            and c['chunk_size_param'] == chunk_size 
            and c['overlap_param'] == overlap
        ]
        
        df_data = []
        for chunk in filtered_chunks:
            df_data.append({
                'Chunk ID': chunk['chunk_id'],
                'Source URL': chunk['source_url'],
                'Page Title': chunk['page_title'],
                'Position in Doc': chunk['position_in_doc'],
                'Character Count': chunk['char_count'],
                'Word Count': chunk['word_count'],
                'Chunk Text': chunk['chunk_text'],
            })
        
        return pd.DataFrame(df_data)
    
    def create_comparison_sheet(self) -> pd.DataFrame:
        """Create side-by-side comparison of first chunk from each strategy"""
        # Get first chunk from each unique strategy configuration
        strategy_configs = {}
        for chunk in self.chunks:
            key = f"{chunk['strategy']}_{chunk['chunk_size_param']}_{chunk['overlap_param']}"
            if key not in strategy_configs:
                strategy_configs[key] = chunk
        
        comparison_data = []
        for key, chunk in strategy_configs.items():
            comparison_data.append({
                'Strategy': chunk['strategy'],
                'Parameters': f"{chunk['chunk_size_param']}/{chunk['overlap_param']}",
                'Char Count': chunk['char_count'],
                'Word Count': chunk['word_count'],
                'Sample Text (First 500 chars)': chunk['chunk_text'][:500] + '...' if len(chunk['chunk_text']) > 500 else chunk['chunk_text'],
            })
        
        return pd.DataFrame(comparison_data)
    
    def export(self):
        """Export all data to Excel with formatting"""
        print(f"\nExporting to {self.output_file}...")
        
        # Create Excel writer
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Summary Sheet
            print("  Creating Summary sheet...")
            df_summary = self.create_summary_sheet()
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # All Chunks Sheet
            print("  Creating All Chunks sheet...")
            df_all_chunks = self.create_all_chunks_sheet()
            df_all_chunks.to_excel(writer, sheet_name='All Chunks', index=False)
            
            # Strategy Comparison Sheet
            print("  Creating Strategy Comparison sheet...")
            df_comparison = self.create_comparison_sheet()
            df_comparison.to_excel(writer, sheet_name='Strategy Comparison', index=False)
            
            # Individual Strategy Sheets
            strategies = [
                ('RecursiveCharacterTextSplitter', 1000, 200, 'Recursive 1000_200'),
                ('RecursiveCharacterTextSplitter', 500, 100, 'Recursive 500_100'),
                ('CharacterTextSplitter', 1000, 200, 'Character 1000_200'),
                ('MarkdownTextSplitter', 1000, 200, 'Markdown 1000_200'),
            ]
            
            for strategy_name, chunk_size, overlap, sheet_name in strategies:
                print(f"  Creating {sheet_name} sheet...")
                df_strategy = self.create_strategy_sheet(strategy_name, chunk_size, overlap)
                if not df_strategy.empty:
                    df_strategy.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        print("  Applying formatting...")
        self._apply_formatting()
        
        print(f"✅ Export complete: {self.output_file}")
    
    def _apply_formatting(self):
        """Apply Excel formatting to all sheets"""
        wb = load_workbook(self.output_file)
        
        # Colors for different sheets
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set reasonable limits
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Enable filters
            ws.auto_filter.ref = ws.dimensions
        
        wb.save(self.output_file)


if __name__ == "__main__":
    # This is just for testing - normally called from main.py
    print("This module should be imported and used with RAGChunker")
